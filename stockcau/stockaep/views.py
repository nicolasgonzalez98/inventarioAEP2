from django.shortcuts import render, redirect, reverse
from django.contrib import messages
from django.contrib.auth.models import User, auth
from .models import *
from .forms import *
from .decorators import *
import openpyxl
from django.contrib.auth.models import Group
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from .filters import HardwareFilter, AsignacionFilter
from django.db.models import F
import os




from django.core.paginator import EmptyPage, Paginator
import subprocess

PRODUCTS_PER_PAGE = 25


##Funciones

def comprimir_archivos_rar(archivos, nombre_archivo_rar):
    # Crea el comando para comprimir los archivos en un archivo RAR
    comando = ['rar', 'a', nombre_archivo_rar] + archivos
    
    # Ejecuta el comando utilizando subprocess
    subprocess.run(comando)

def mayus_minus(pal):
    if(pal == ''):
        return pal
    pal.strip()
    if pal[-1] == ' ':
        pal.replace(' ', '')
    return pal.lower().capitalize()

def buscar_repetido(nro_de_serie):
    try:
        hard = Hardware.objects.get(nro_de_serie = nro_de_serie)
    except:
        hard = None
    return hard
# Create your views here.

@login_required(login_url='login')
def index(request):
    cant_pags = []
    page = request.GET.get('page',1)
    asignacion = request.GET.get('asignacion', False)
    editar = request.GET.get('editar', False)
    agregar = request.GET.get('agregar', False)
    notificaciones = Notificacion.objects.filter(realizado = True)
    notificaciones = len(notificaciones)
    
    
    f = HardwareFilter(request.GET, queryset=Hardware.objects.all())
    
    product_paginator = Paginator(list(f.qs), PRODUCTS_PER_PAGE)
    
    try:
        pagina = product_paginator.page(page)
    except EmptyPage:
        pagina = product_paginator.page(product_paginator.num_pages)
    except:
        pagina = product_paginator.page(product_paginator.num_pages)

    for i in list(product_paginator.page_range):
        if i <= int(page) + 5 and i>=int(page)-5:
            cant_pags.append(i)

    ctx = {
        'link':'index',
        'filter':f,
        'pagina': pagina,
        'paginator':product_paginator,
        'cant_pags':cant_pags,
        'num_page':int(page),
        'cant_notificaciones':notificaciones
    }

    if asignacion == '1':
        ctx['asignacion'] = True
    if editar == '1':
        ctx['editar'] = True
    if agregar == '1':
        ctx['agregar'] = True
    
    
    return render(request, 'main.html', ctx)

@unauthorized_user
def register(request):
    if request.method == 'POST':
        username = request.POST['username']
        email = request.POST['email']
        password = request.POST['password']
        password2 = request.POST['password2']
        name = request.POST['name']
        surname = request.POST['surname']
        
        

        if password == password2:
            if User.objects.filter(email=email).exists():
                messages.info(request, 'Email ya usado')
                return redirect('register')
            elif User.objects.filter(username=username).exists():
                messages.info(request, 'Username Taken')
                return redirect('register')
            else:
                user = User.objects.create_user(username=username, email=email, password=password)
                user.save()
                group, create=Group.objects.get_or_create(name='user')
                user_model = User.objects.get(username=username)
                user_model.groups.add(group)
                nuevo_tecnico = Tecnico.objects.create(user = user_model, id_user = user_model.id, nombre=name, apellido = surname)
                nuevo_tecnico.save()
                return redirect('login')
        else:
            messages.info(request, 'Las contraseñas no coinciden.')
            return redirect('register') 
    return render(request,'signup.html')

@unauthorized_user
def login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        

        user = auth.authenticate(username=username, password=password)
        

        if user is not None:
            auth.login(request, user)
            return redirect('/')
        else:
            messages.info(request,'Credentials invalid')
            return redirect('login')
    return render(request, 'signin.html')

@login_required(login_url='login')
def logout(request):
    auth.logout(request)
    return redirect('login')

@login_required(login_url='login')
def add_inventary(request):
    
    form_add_inventary = HardwareForm()
    
    ctx = {'link':'create'}
    ctx['form_add_inventary'] = form_add_inventary

    if request.method == 'POST':
        form = HardwareForm(request.POST)
        if form.is_valid():
            try:
                form.cleaned_data['nro_de_serie'] = form.cleaned_data['nro_de_serie'].upper()
            except:
                pass
            if len(form.cleaned_data['nro_de_serie']) > 0 and ('?' not in form.cleaned_data['nro_de_serie']):
                hard = buscar_repetido(form.cleaned_data['nro_de_serie'])
            else:
                hard = None
            
            if hard != None:
                messages.info(request, 'Ya hay un hardware en el inventario con el mismo numero de serie.')
                ctx['form_add_inventary'] = HardwareForm(form.cleaned_data)
            else:
                
                if len(form.cleaned_data['nro_de_serie']) == 0 or ('?' in form.cleaned_data['nro_de_serie']):
                    form.cleaned_data['nro_de_serie'] = 'S/D'
                tipo, create = Tipo.objects.get_or_create(id = request.POST['tipo'])
                marca, create = Marca.objects.get_or_create(id = request.POST['marca']) 
                modelo = Modelo.objects.get(id = request.POST['modelo'])
                ubicacion, create = Ubicacion.objects.get_or_create(id=request.POST['ubicacion'])
                estado, create = Estado.objects.get_or_create(id = request.POST['estado'])
                

                hardware = Hardware.objects.create(tipo = tipo, marca=marca, modelo=modelo,ubicacion=ubicacion, estado = estado, nro_de_serie=form.cleaned_data['nro_de_serie'], observaciones = request.POST['observaciones'], origen = request.POST['origen'], nota = request.POST['nota'])
                
                if(request.user.is_staff == False):
                    Notificacion.objects.create(hardware=hardware, usuario = request.user, tipo = 'CREATE')
                
                return redirect(reverse('index')+f'?agregar=1')
    return render(request, 'main.html', ctx)

@login_required(login_url='login')
@admin_only
def reload(request):
    # df = openpyxl.load_workbook("INVENTARIO_YENNY.xlsx")
    # print('yenny')
    # dataframe = df.active
    # data = []
    # for row in range(1, dataframe.max_row):
    #     _row=[row]
    #     for col in dataframe.iter_cols(1,dataframe.max_column):
    #         _row.append(col[row].value)
    #     data.append(_row)

    # for dato in data:
    #         print(dato)
    #         tipo, create = Tipo.objects.get_or_create(name = "Pantalla")
    #         marca, create = Marca.objects.get_or_create(nombre = mayus_minus(str(dato[1]))) 
    #         modelo, create = Modelo.objects.get_or_create(nombre = mayus_minus(str(dato[2])))
    #         ubicacion, create = Ubicacion.objects.get_or_create(nombre=mayus_minus(str(dato[6])))
    #         estado, create = Estado.objects.get_or_create(nombre = dato[5])

    #         if dato[7] == None:
    #             dato[7] = ''
    #         else:
    #             dato[7] = mayus_minus(str(dato[7]))

    #         hard = Hardware.objects.create(tipo=tipo, marca=marca, modelo=modelo, ubicacion=ubicacion, estado = estado, nro_de_serie=mayus_minus(str(dato[4])).upper(), observaciones = dato[7], origen = 'Yenny')
    #         hard.save()

    # df = openpyxl.load_workbook("Mueble Cau.xlsx")
    # print('mueble cau')
    
    # for i in df.sheetnames:
        
    #     if df[i].title == 'Asignaciones':
    #         pass
    #     else:
    #         dataframe = df[i]
    #         data = []
    #         for row in range(1, dataframe.max_row):
    #             _row=[row]
    #             for col in dataframe.iter_cols(1,dataframe.max_column):
    #                 _row.append(col[row].value)
    #             data.append(_row)
            
    #         for dato in data:
                    
    #                 tipo, create = Tipo.objects.get_or_create(name = mayus_minus(str(dato[1])))
    #                 marca, create = Marca.objects.get_or_create(nombre = mayus_minus(str(dato[2]))) 
    #                 modelo, create = Modelo.objects.get_or_create(nombre = mayus_minus(str(dato[3])))
    #                 ubicacion, create = Ubicacion.objects.get_or_create(nombre=dato[6])
    #                 estado, create = Estado.objects.get_or_create(nombre = dato[5])

    #                 if dato[7] == None:
    #                     dato[7] = ''
    #                 else:
    #                     dato[7] = mayus_minus(str(dato[7]))
    #                 if dato[4] == None:
    #                     dato[4] = 'S/D'
                    
    #                 try:
    #                     if dato[8] == None:
    #                         dato[8] = ""

    #                     hard = Hardware.objects.create(tipo=tipo, marca=marca, modelo=modelo, ubicacion=ubicacion, estado = estado, nro_de_serie=mayus_minus(str(dato[4])).upper(), observaciones = dato[7], nota=str(dato[8]))
    #                 except:
    #                     hard = Hardware.objects.create(tipo=tipo, marca=marca, modelo=modelo, ubicacion=ubicacion, estado = estado, nro_de_serie=mayus_minus(str(dato[4])).upper(), observaciones = dato[7])
    #                 hard.save()

    # df = openpyxl.load_workbook("inventariot4.xlsx")
    
    # for i in df.sheetnames:
    #     if df[i].title == 'Scrap':
    #         pass
    #     else:
    #         dataframe = df[i]
    #         data = []
            
    #         for row in range(1, dataframe.max_row):
    #             _row=[row]
    #             for col in dataframe.iter_cols(1,dataframe.max_column):
    #                 _row.append(col[row].value)
    #             data.append(_row)

    #         for dato in data:
                
    #             tipo, create = Tipo.objects.get_or_create(name = mayus_minus(str(dato[1])))
    #             marca, create = Marca.objects.get_or_create(nombre = mayus_minus(str(dato[2]))) 
    #             modelo, create = Modelo.objects.get_or_create(nombre = mayus_minus(str(dato[3])))
    #             ubicacion, create = Ubicacion.objects.get_or_create(nombre=mayus_minus(str(dato[6])))
    #             estado, create = Estado.objects.get_or_create(nombre = dato[5])

    #             if dato[7] == None:
    #                 dato[7] = ''
    #             else:
    #                 if len(dato) > 8:
    #                     dato[7] = mayus_minus(str(dato[7]))
    #                 else:
    #                     dato[7] = mayus_minus(str(dato[7]))
    #             try:
    #                 if dato[8] == None:
    #                     dato[8] = ""
    #                 hard = Hardware.objects.create(tipo=tipo, marca=marca, modelo=modelo, ubicacion=ubicacion, estado = estado, nro_de_serie=mayus_minus(str(dato[4])).upper(), observaciones = dato[7], origen = "T4", nota=str(dato[8]))
    #             except:
    #                 hard = Hardware.objects.create(tipo=tipo, marca=marca, modelo=modelo, ubicacion=ubicacion, estado = estado, nro_de_serie=mayus_minus(str(dato[4])).upper(), observaciones = dato[7], origen = "T4")
    #             hard.save()

    df = openpyxl.load_workbook("Mueble Cau.xlsx")
    for i in df.sheetnames:
        if df[i].title == 'Asignaciones':
            dataframe = df[i]
            data = []
            for row in range(1, dataframe.max_row):
                _row=[row]
                for col in dataframe.iter_cols(1,dataframe.max_column):
                    _row.append(col[row].value)
                data.append(_row)
            
            for dato in data:
                
                try:
                    if (dato[3] == None):
                        hard = Hardware.objects.get(nro_de_serie = dato[4])
                        
                    else:
                        hard = Hardware.objects.get(nro_de_serie = dato[3])
                        
                    if dato[11] == None:
                            dato[11] = ""
                    if dato[9] == None:
                            dato[9] = ""
                    
                    
                    asignacion = Asignacion.objects.create(hardware = hard, usuario = dato[7], fecha_creacion = dato[10].date(), nro_ticket = dato[11], nota = dato[9])
                    asignacion.save()
                except:
                    print(dato)

    return redirect('index')

@login_required(login_url='login')
def delete(request, id):
    if(request.user.is_staff):
        hardware = Hardware.objects.get(id=id)
        hardware.delete()
        return redirect('index')
    else:
        hardware = Hardware.objects.get(id=id)
        usuario = User.objects.get(username = request.user)
        Notificacion.objects.create(hardware=hardware, usuario = usuario, tipo = 'DELETE')
        return redirect('index')

@login_required(login_url='login')
def edit(request, id):
    try:
        to_edit = Hardware.objects.get(id=id)
    except:
        return redirect('index')
    ctx = {}
    ctx['to_edit'] = to_edit
    edit_form = HardwareForm(to_edit.toJSON())
    ctx['edit_form'] = edit_form
    ctx['link'] = 'edit'
    
    
    if request.method == 'POST':
        
        if request.POST['nro_de_serie'].upper() != to_edit.nro_de_serie:
            
            if len(request.POST['nro_de_serie']) > 0 and ('?' not in request.POST['nro_de_serie']):
                hard = Hardware.objects.filter(nro_de_serie = request.POST['nro_de_serie'].upper())
                
            else:
                hard = False

            if hard:
                messages.info(request, 'Ya hay un hardware en el inventario con el mismo numero de serie.')
                ctx['form_add_inventary'] = HardwareForm(request.POST)
                return render(request, 'main.html', ctx)
        
        to_edit.tipo  = Tipo.objects.get(id=request.POST['tipo'])
        to_edit.marca = Marca.objects.get(id=request.POST['marca'])
        to_edit.modelo = Modelo.objects.get(id=request.POST['modelo'])
        to_edit.ubicacion = Ubicacion.objects.get(id=request.POST['ubicacion'])
        to_edit.observaciones = request.POST['observaciones']
        to_edit.origen = request.POST['origen']
        to_edit.nota = request.POST["nota"]
        if(request.user.is_staff):

            to_edit.nro_de_serie = request.POST['nro_de_serie'].upper()
            to_edit.estado = Estado.objects.get(id=request.POST['estado'])
        else:
            nro_serie = ''
            estado = ''
            if to_edit.nro_de_serie != request.POST['nro_de_serie']:
                nro_serie= request.POST['nro_de_serie'].upper()
            if to_edit.estado != Estado.objects.get(id=request.POST['estado']):
                estado= (Estado.objects.get(id=request.POST['estado'])).nombre
    
            Notificacion.objects.create(hardware=to_edit, usuario = User.objects.get(username = request.user), tipo = 'EDIT', nro_de_serie = nro_serie, estado = estado)
        to_edit.save()
        
        return redirect(reverse('index')+f'?editar=1')
    
    return render(request, 'main.html', ctx)

@admin_only
@login_required(login_url='login')
def get_info(request):
    data = list(Hardware.objects.values())
    
    return JsonResponse(data, safe=False)

@login_required(login_url='login')
@admin_only
def notificaciones(request):
    ctx={'link':'notification'}
    status = request.GET.get('status', False)
    

    if status == 'cancel':
        ctx['status'] = True
        ctx['title'] = 'Peticion cancelada'
        ctx['msg'] = 'Se ha cancelado la petición correctamente.'
    elif status == 'accept':
        ctx['status'] = True
        ctx['title'] = 'Peticion aprobada'
        ctx['msg'] = 'Se ha aprobado la petición correctamente.'
    
    notificaciones = Notificacion.objects.filter(realizado = False)
    
    ctx['notificaciones'] = notificaciones
    ctx['cant_notificaciones'] = len(notificaciones)
    
    return render(request, 'main.html', ctx)

@login_required(login_url='login')
@admin_only
def accion_notificacion(request):
    id = request.GET.get('id')
    status = request.GET.get('status')
    notificacion = Notificacion.objects.get(id=id)

    
    if (notificacion.tipo == 'CREATE' and status == 'cancel') or (notificacion.tipo == 'DELETE' and status == 'accept'):
        hardware = Hardware.objects.get(id = notificacion.hardware.id)
        hardware.delete()
    elif(notificacion.tipo == 'EDIT' and status == 'accept'):
        hardware = Hardware.objects.get(id = notificacion.hardware.id)
        if notificacion.nro_de_serie:
            hardware.nro_de_serie = notificacion.nro_de_serie
        if notificacion.estado:
            hardware.estado = Estado.objects.get(nombre = notificacion.estado)
        hardware.save()

    notificacion.realizado = True
    notificacion.save()

    return redirect(reverse('notifications') + f'?status={status}')

@login_required(login_url='login')
def asignacion(request):
    id = request.GET.get('id')
    hardware = Hardware.objects.get(id=id)
    if request.method == 'POST':
        person = request.POST['person']
        nro_ticket = request.POST['nro_ticket']
        if len(person) == 0:
            return redirect('index')
        if nro_ticket == "":
            Asignacion.objects.create(hardware=hardware, usuario=person)
        else:
            Asignacion.objects.create(hardware=hardware, usuario=person, nro_ticket = nro_ticket)
        return redirect(reverse('index')+f'?asignacion=1')
    
    return redirect('index')

@login_required(login_url='login')
def asignaciones(request):
    ctx = {'link':'asignaciones'}
    asignaciones = Asignacion.objects.filter()
    f = AsignacionFilter(request.GET, queryset=Asignacion.objects.all())
    asignaciones_filtradas = f.qs
    ctx['filter'] = f
    
    ctx['asignaciones'] = asignaciones_filtradas
    return render(request, 'main.html', ctx)

@login_required(login_url='login')
@admin_only
def administrar_users(request):
    usuarios = Tecnico.objects.all()
    ctx = {}
    ctx["users"] = usuarios
    ctx["link"] = "admin_users"
    return render(request, "main.html", ctx)

@login_required(login_url='login')
@admin_only
def to_admin(request):
    id = request.GET.get('id')
    status = request.GET.get("status")

    if status == "1":
        group=Group.objects.get(name='admin')
        user = User.objects.get(id=id)
        user.is_superuser = True
        user.is_staff = True
        user.groups.add(group)
        group=Group.objects.get(name='user')
        user.groups.remove(group)
    elif status == "0":
        group=Group.objects.get(name='user')
        user = User.objects.get(id=id)
        user.is_superuser = False
        user.is_staff = False
        user.groups.add(group)
        group=Group.objects.get(name='admin')
        user.groups.remove(group)
    user.save()

    return redirect("admin_users")

@login_required(login_url='login')
@admin_only
def to_active(request):
    id = request.GET.get('id')
    status = request.GET.get("status")

    if status == "1":
        user = User.objects.get(id=id)
        user.is_active = True
    elif status == "0":
        user = User.objects.get(id=id)
        user.is_active = False
    user.save()
    return redirect("admin_users")

@login_required(login_url='login')
def importar_datos(request):
    hojas = ["Stock","Asignaciones"]
    separador = os.path.sep
    dir_actual = os.path.dirname(os.path.abspath(__file__))
    direc = separador.join(dir_actual.split(separador)[:-1])
    directorio_static = os.path.join(direc, 'static')
    nombre_archivo = "nuevo_inventario.xlsx"
    
    asignaciones = Asignacion.objects.all().values("hardware__tipo__name", "hardware__marca__nombre","hardware__nro_de_serie","hardware__modelo__nombre","hardware__estado__nombre",'usuario',"hardware__ubicacion__nombre", 'nota','fecha_creacion', 'nro_ticket')
    # for asignacion in asignaciones:
    #     asignacion['fecha_creacion_sin_tz'] = asignacion['fecha_creacion_sin_tz'].replace(tzinfo=None)
    hardware = Hardware.objects.all().values('tipo__name', 'marca__nombre', 'modelo__nombre', 'nro_de_serie', 'estado__nombre', 'ubicacion__nombre', 'observaciones', 'nota')
    
    wb = openpyxl.Workbook(nombre_archivo)
    
    for i in range(len(hojas)):
        wb.create_sheet(hojas[i])
        hoja = wb[hojas[i]]
        wb.active = hoja
        hoja = wb.active
        if hoja.title in ["Asignaciones"]:
            hoja.append(("Producto","Marca","Service Tag / IMEI", "Modelo",	"Estado","Usuario",	"Ubicación","Nota","N° Ticket",	"Fecha"))
            for h in asignaciones:
                fila_limpia = tuple(str(valor).replace("None", "") if "None" in str(valor) else valor for valor in h.values())
                
                hoja.append(fila_limpia)
        else: 
            hoja.append(("Tipo", "Marca", "Modelo", "Service Tag/IMEI", "Estado", "Ubicación","Observaciones", "Nota" ))
            for h in hardware:
                fila_limpia = tuple(str(valor).replace("None", "") if "None" in str(valor) else valor for valor in h.values())
                hoja.append(fila_limpia)
        
        
                
    # Ruta completa del archivo Excel en la carpeta 'static'
    ruta_archivo_excel = os.path.join(directorio_static, nombre_archivo)
    # Guardar el libro en la carpeta 'static'
    wb.save(ruta_archivo_excel)
    # Cerrar el libro después de guardarlo
    wb.close()
    
    return render(request, 'main.html', {"link":"realizar_informe"})

@login_required(login_url='login')
def realizar_informes(request):
    return render(request, 'main.html', {"link":"realizar_informe"})

@login_required(login_url='login')
def cambio_contraseña(request):
    if request.method == 'POST':
        form = CambioContraseñaForm(request.POST)
        if form.is_valid():
            nueva_contraseña = form.cleaned_data['nueva_contraseña']
            request.user.set_password(nueva_contraseña)
            request.user.save()
            return redirect('index')
        else:
            form.add_error(None, "Las contraseñas no coinciden.")
            
            return render(request, 'main.html', {'form': form, "link":"cambio_contraseña"})# Redirigir a la página de perfil o cualquier otra página después de cambiar la contraseña
    else:
        form = CambioContraseñaForm()

    return render(request, 'main.html', {'form': form, "link":"cambio_contraseña"})